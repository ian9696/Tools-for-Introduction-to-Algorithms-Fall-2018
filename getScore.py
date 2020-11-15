import requests
import json
import os
import shutil
import datetime
import re
import openpyxl

with open('token.txt') as f:
	COOKIES={'token': f.read().rstrip()}

LOG_PATH='log/'

PROBLEMS=list(range(754, 760+1))

def get(url, printR=True, saveR=True):
    if printR:
        print(url)
    r=ses.get(url)
    if saveR:
        with open(os.path.join(LOG_PATH, re.sub(r'[:?/]', '-', url)+'.json'), 'wb') as f:
            f.write(r.content)
    if printR:
        print('r.status_code='+str(r.status_code))
        print('='*30)
    if r.status_code!=200:
        raise
    return r

def saveLog(x, filename):
    filename=os.path.join(LOG_PATH, filename+'.json')
    with open(filename, 'w', newline='\n', encoding='utf-8') as f:  
            json.dump(x, f, indent=' '*4)

if os.path.exists(LOG_PATH):
    shutil.rmtree(LOG_PATH)
os.mkdir(LOG_PATH)

with requests.Session() as ses:
    for k, v in COOKIES.items():
        ses.cookies.set(k, v)
    
    r=get('https://api.oj.nctu.me/submissions/?count=1000000&page=1&group_id=19')
    subs=json.loads(r.content)['msg']['submissions']
    saveLog(subs, 'subs')
    
    l=len(subs)
    print(str(l)+' submissions')
    subs=[sub for sub in subs if isinstance(sub['score'], int) and sub['score'] in range(0, 101)]
    if len(subs)!=l:
        print('warning: '+str(l-len(subs))+' submissions have invalid scores')
        print(str(len(subs))+' submissions have valid scores')
    else:
        print('all submissions have valid scores')
    print('='*30)
    saveLog(subs, 'subs, valid score')
    
    subs,l=[],subs
    ds=datetime.datetime.strptime('2019-01-12 13:10:00', '%Y-%m-%d %H:%M:%S')
    dt=datetime.datetime.strptime('2019-01-12 18:10:00', '%Y-%m-%d %H:%M:%S')
    for sub in l:
        d=datetime.datetime.strptime(sub['created_at'], '%Y-%m-%d %H:%M:%S')
        if d>=ds and d<=dt:
            subs+=[sub]
    print(str(len(subs))+' valid score&in time submissions')
    print('='*30)
    saveLog(subs, 'subs, valid score, in time')
    
    r=get('https://api.oj.nctu.me/groups/19/users/')
    usrs=json.loads(r.content)['msg']
    saveLog(usrs, 'usrs')
    
    for i in range(len(subs)):
        print('\r{}/{}'.format(i, len(subs)), end='')
        r=get('https://api.oj.nctu.me/submissions/{}/'.format(subs[i]['id']), False)
        subd=json.loads(r.content)['msg']
        usr=[usr for usr in usrs if subd['user_id']==usr['id']]
        if len(usr)!=1:
            raise
        usr=usr[0]
        subd['user_name']=usr['name']
        subs[i]=subd
    print('\r{}/{}'.format(len(subs), len(subs)))
    print(str(len(subs))+' valid score&in time detailed submissions')
    print('='*30)
    saveLog(subs, 'detailed subs, valid score, in time')
    
    with open('ignore_submission_id.json') as f:
        ignore_submission_id=json.load(f)
    with open('ignore_user_id.json') as f:
        ignore_user_id=json.load(f)
    with open('ignore_user_name.json') as f:
        ignore_user_name=json.load(f)
    with open('allow_ip.json') as f:
        allow_ip=json.load(f)
    subs,l=[],subs
    for sub in l:
        if sub['id'] in ignore_submission_id:
            continue
        if sub['user_id'] in ignore_user_id:
            continue
        if sub['user_name'] in ignore_user_name:
            continue
        if sub['ip'] not in allow_ip:
            continue
        subs+=[sub]
    print(str(len(subs))+' valid score&in time&not ignored&allowed ip detailed submissions')
    print('='*30)
    saveLog(subs, 'detailed subs, valid score, in time, not ignored, allowed ip')
    
    wb=openpyxl.load_workbook('Grades.xlsx')
    ws=wb.active
    for i in range(2, ws.max_row+1):
        for j in range(len(PROBLEMS)):
            user_name=ws['B'+str(i)].value
            pid=PROBLEMS[j]
            sco=0
            for sub in subs:
                if sub['user_name']==user_name and sub['problem_id']==pid:
                    sco=max(sco, sub['score'])
            ws[chr(ord('C')+j)+str(i)]=sco
    wb.save('Grades output.xlsx')
    print('xlsx saved')
    print('='*30)
